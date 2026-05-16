import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

folder_path = "c:/Users/bisra/Music/Task 2/"

df = pd.read_csv(folder_path + 'cleaned_ravenstack_data.csv')

total_accounts = len(df)
total_churned = int(df['is_churned'].sum())
global_churn_rate = float(total_churned / total_accounts)
avg_tickets = float(df['total_tickets'].mean())

plan_churn = df.groupby('plan_tier')['is_churned'].mean().reset_index()
plan_churn.columns = ['Plan Tier', 'Account Churn Rate']

industry_churn = df.groupby('industry')['is_churned'].mean().sort_values(ascending=False).reset_index()
industry_churn.columns = ['Industry Sector', 'Account Churn Rate']

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Executive Summary"
ws.views.sheetView[0].showGridLines = True

COLOR_NAVY = "1E3A8A"       
COLOR_ICE_BLUE = "F0F5FA"   
COLOR_CARD_BG = "F8FAFC"    
COLOR_BORDER = "CBD5E1"     

font_title = Font(name="Segoe UI", size=18, bold=True, color="0F172A")
font_section = Font(name="Segoe UI", size=13, bold=True, color="1E3A8A")
font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
font_regular = Font(name="Segoe UI", size=11, color="334155")
font_bold = Font(name="Segoe UI", size=11, bold=True, color="0F172A")
font_card_val = Font(name="Segoe UI", size=16, bold=True, color="1E3A8A")
font_card_lbl = Font(name="Segoe UI", size=9, bold=False, color="64748B")

fill_header = PatternFill(start_color=COLOR_NAVY, end_color=COLOR_NAVY, fill_type="solid")
fill_zebra = PatternFill(start_color=COLOR_ICE_BLUE, end_color=COLOR_ICE_BLUE, fill_type="solid")
fill_card = PatternFill(start_color=COLOR_CARD_BG, end_color=COLOR_CARD_BG, fill_type="solid")

thin_side = Side(border_style="thin", color=COLOR_BORDER)
border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

ws['A1'] = "RavenStack Customer Retention & Churn Analytics"
ws['A1'].font = font_title
ws.row_dimensions[1].height = 30

def create_kpi_card(sheet, start_col, label, value, number_format):
    col1 = get_column_letter(start_col)
    col2 = get_column_letter(start_col + 1)
    
    sheet.merge_cells(f"{col1}3:{col2}3")
    sheet.merge_cells(f"{col1}4:{col2}4")
    
    lbl_cell = sheet[f"{col1}3"]
    val_cell = sheet[f"{col1}4"]
    
    lbl_cell.value = label
    lbl_cell.font = font_card_lbl
    lbl_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    val_cell.value = value
    val_cell.font = font_card_val
    val_cell.alignment = Alignment(horizontal="center", vertical="center")
    val_cell.number_format = number_format
    
    for r in [3, 4]:
        for c in [start_col, start_col + 1]:
            cell = sheet.cell(row=r, column=c)
            cell.fill = fill_card
            cell.border = border_all

create_kpi_card(ws, 1, "TOTAL CORPORATE ACCOUNTS", total_accounts, "#,##0")
create_kpi_card(ws, 3, "TOTAL LOST ACCOUNTS (CHURN)", total_churned, "#,##0")
create_kpi_card(ws, 5, "GLOBAL PORTFOLIO CHURN RATE", global_churn_rate, "0.00%")
create_kpi_card(ws, 7, "AVG SUPPORT TICKETS / USER", avg_tickets, "0.0")
ws.row_dimensions[3].height = 18
ws.row_dimensions[4].height = 26

def build_polished_table(sheet, start_row, section_title, dataframe):
    sheet.cell(row=start_row, column=1, value=section_title).font = font_section
    sheet.row_dimensions[start_row].height = 24
    
    header_row = start_row + 1
    sheet.row_dimensions[header_row].height = 22
    for c_idx, col_name in enumerate(dataframe.columns, start=1):
        cell = sheet.cell(row=header_row, column=c_idx, value=col_name)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="left" if c_idx == 1 else "right", vertical="center")
        cell.border = border_all
        
    for r_idx, row_values in dataframe.iterrows():
        current_row = header_row + 1 + r_idx
        sheet.row_dimensions[current_row].height = 20
        for c_idx, val in enumerate(row_values, start=1):
            cell = sheet.cell(row=current_row, column=c_idx, value=val)
            cell.font = font_regular
            cell.border = border_all
            
            if r_idx % 2 == 1:
                cell.fill = fill_zebra
                
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if "Rate" in dataframe.columns[c_idx-1]:
                    cell.number_format = "0.00%"
                else:
                    cell.number_format = "#,##0"
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

build_polished_table(ws, 7, "1. Retention Distribution by Subscription Tier", plan_churn)
build_polished_table(ws, 14, "2. Churn Concentrations across Industry Sectors", industry_churn)

for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            val_str = str(cell.value)
            if len(val_str) < 50:
                max_len = max(max_len, len(val_str))
    ws.column_dimensions[col_letter].width = max(max_len + 5, 14)

output_path = folder_path + "ravenstack_retention_report_v2.xlsx"
wb.save(output_path)
print("🏁 Success! The beautifully redesigned workbook has been saved.")